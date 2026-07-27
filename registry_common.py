#!/usr/bin/env python3
"""
registry_common.py
==================
Shared helpers for all registry fetchers: HTTP with retries, generic
JSON/XML/HTML flatteners, the unified output schema, and Excel writing.

No LLM / AI service is used anywhere in this package.
"""

from __future__ import annotations

import html as html_mod
import json
import os
import re
import sys
import time
from typing import Any, Dict, Iterable, List, Optional

import requests

# ── Canonical source names (the ONLY values allowed in the `source` column) ──
SRC_EUCT = "EU Clinical Trials"
SRC_CTGOV = "ClinicalTrials.gov"
SRC_CTRI = "CTRI"
SRC_CHICTR = "ChiCTR"
SRC_JRCT = "JRCT"
SRC_ANZCTR = "ANZCTR"
SRC_CRIS = "CRIS"
SRC_REBEC = "ReBEC"

ALLOWED_SOURCES = [
    SRC_EUCT, SRC_CTGOV, SRC_CTRI, SRC_CHICTR,
    SRC_JRCT, SRC_ANZCTR, SRC_CRIS, SRC_REBEC,
]

# ── Unified schema shared by every registry ──────────────────────────────────
UNIFIED_COLUMNS = [
    "source",
    "trial_id",
    "secondary_ids",
    "title",
    "public_title",
    "status",
    "phase",
    "study_type",
    "study_design",
    "conditions",
    "interventions",
    "drug_names",
    "sponsor",
    "sponsor_type",
    "collaborators",
    "countries",
    "sites",
    "target_enrollment",
    "actual_enrollment",
    "age_min",
    "age_max",
    "gender",
    "healthy_volunteers",
    "inclusion_criteria",
    "exclusion_criteria",
    "primary_objective",
    "primary_outcome",
    "secondary_outcome",
    "start_date",
    "completion_date",
    "registration_date",
    "last_updated",
    "results_available",
    "findings",
    "contact",
    "ethics_approval",
    "url",
]

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/125.0 Safari/537.36"
)

DEFAULT_TIMEOUT = 60
DEFAULT_RETRIES = 3
RETRY_BACKOFF = 3


# ── HTTP ─────────────────────────────────────────────────────────────────────
def make_session(extra_headers: Optional[Dict[str, str]] = None) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept-Language": "en,en-US;q=0.9",
    })
    if extra_headers:
        s.headers.update(extra_headers)
    return s


def http_get(session: requests.Session, url: str, *, params=None,
             timeout: int = DEFAULT_TIMEOUT, retries: int = DEFAULT_RETRIES,
             expect_json: bool = False, **kw):
    """GET with retry/backoff. Returns text, or parsed JSON if expect_json."""
    last = None
    for attempt in range(1, retries + 1):
        try:
            r = session.get(url, params=params, timeout=timeout, **kw)
            if r.status_code in (429, 500, 502, 503, 504):
                raise requests.HTTPError(f"HTTP {r.status_code}", response=r)
            r.raise_for_status()
            return r.json() if expect_json else r.text
        except Exception as exc:
            last = exc
            if attempt < retries:
                time.sleep(RETRY_BACKOFF * attempt)
    raise RuntimeError(f"GET {url} failed after {retries} attempts: {last}")


def http_post(session: requests.Session, url: str, *, data=None, json_body=None,
              timeout: int = DEFAULT_TIMEOUT, retries: int = DEFAULT_RETRIES,
              expect_json: bool = False, **kw):
    last = None
    for attempt in range(1, retries + 1):
        try:
            r = session.post(url, data=data, json=json_body, timeout=timeout, **kw)
            if r.status_code in (429, 500, 502, 503, 504):
                raise requests.HTTPError(f"HTTP {r.status_code}", response=r)
            r.raise_for_status()
            return r.json() if expect_json else r.text
        except Exception as exc:
            last = exc
            if attempt < retries:
                time.sleep(RETRY_BACKOFF * attempt)
    raise RuntimeError(f"POST {url} failed after {retries} attempts: {last}")


# ── Text helpers ─────────────────────────────────────────────────────────────
def clean(text: Any) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", html_mod.unescape(str(text))).strip()


def join(values: Iterable[Any], sep: str = "; ") -> str:
    out, seen = [], set()
    for v in values or []:
        s = clean(v)
        if s and s not in seen:
            seen.add(s)
            out.append(s)
    return sep.join(out)


def first_nonempty(*values) -> str:
    for v in values:
        s = clean(v)
        if s:
            return s
    return ""


# ── Generic flatteners: these are how we capture "all available fields" ──────
def flatten_json(obj: Any, prefix: str = "", out: Optional[Dict[str, str]] = None,
                 max_list: int = 50) -> Dict[str, str]:
    """
    Recursively flatten arbitrary JSON into dotted keys.
    Lists of scalars collapse to a '; '-joined string; lists of objects are
    indexed (key.0.field, key.1.field, ...).
    """
    if out is None:
        out = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            flatten_json(v, key, out, max_list)
    elif isinstance(obj, list):
        if not obj:
            return out
        if all(not isinstance(x, (dict, list)) for x in obj):
            out[prefix] = join(obj)
        else:
            for i, item in enumerate(obj[:max_list]):
                flatten_json(item, f"{prefix}.{i}", out, max_list)
            if len(obj) > max_list:
                out[f"{prefix}._truncated"] = str(len(obj))
    else:
        val = clean(obj)
        if val:
            out[prefix] = val
    return out


def flatten_xml(element, prefix: str = "", out: Optional[Dict[str, str]] = None
                ) -> Dict[str, str]:
    """Recursively flatten an ElementTree element into dotted keys."""
    if out is None:
        out = {}
    tag = element.tag.split("}")[-1]
    key = f"{prefix}.{tag}" if prefix else tag

    for attr, val in (element.attrib or {}).items():
        akey = f"{key}@{attr}"
        if clean(val):
            out[akey] = clean(val)

    children = list(element)
    if children:
        counts: Dict[str, int] = {}
        for child in children:
            ctag = child.tag.split("}")[-1]
            counts[ctag] = counts.get(ctag, 0) + 1
        seen: Dict[str, int] = {}
        for child in children:
            ctag = child.tag.split("}")[-1]
            if counts[ctag] > 1:
                idx = seen.get(ctag, 0)
                seen[ctag] = idx + 1
                flatten_xml(child, f"{key}.{idx}", out)
            else:
                flatten_xml(child, key, out)
    else:
        text = clean(element.text)
        if text:
            if key in out:
                out[key] = out[key] + "; " + text
            else:
                out[key] = text
    return out


def extract_label_value_pairs(soup) -> Dict[str, str]:
    """
    Generic HTML field harvester: pulls every label/value pair it can find from
    two-column tables, definition lists, and label/value div patterns.
    Used by the HTML-scraped registries to capture all displayed fields.
    """
    pairs: Dict[str, str] = {}

    def add(label: str, value: str):
        label = clean(label).rstrip(":").strip()
        value = clean(value)
        if not label or not value or len(label) > 150:
            return
        if label.lower() == value.lower():
            return
        if label in pairs:
            if value not in pairs[label]:
                pairs[label] = pairs[label] + " | " + value
        else:
            pairs[label] = value

    # two-column (or label + last-cell) tables
    for tr in soup.find_all("tr"):
        cells = tr.find_all(["td", "th"], recursive=False) or tr.find_all(["td", "th"])
        if len(cells) == 2:
            add(cells[0].get_text(" ", strip=True), cells[1].get_text(" ", strip=True))
        elif len(cells) == 3:
            # code | label | value  (EudraCT / CTRI style)
            add(cells[1].get_text(" ", strip=True), cells[2].get_text(" ", strip=True))

    # definition lists
    for dl in soup.find_all("dl"):
        dts, dds = dl.find_all("dt"), dl.find_all("dd")
        for dt, dd in zip(dts, dds):
            add(dt.get_text(" ", strip=True), dd.get_text(" ", strip=True))

    # label/value div or span pairs
    for tag in soup.find_all(["div", "li", "p"]):
        label_el = tag.find(["label", "strong", "b", "th"], recursive=False)
        if label_el:
            label = label_el.get_text(" ", strip=True)
            full = tag.get_text(" ", strip=True)
            if full.startswith(label) and len(full) > len(label):
                add(label, full[len(label):])

    return pairs


def find_field(pairs: Dict[str, str], *needles: str, exact: bool = False) -> str:
    """Look up a harvested field by fuzzy label match (case-insensitive)."""
    lowered = {k.lower(): v for k, v in pairs.items()}
    for needle in needles:
        n = needle.lower()
        if exact:
            if n in lowered:
                return lowered[n]
            continue
        for k, v in lowered.items():
            if n == k:
                return v
    for needle in needles:
        n = needle.lower()
        for k, v in lowered.items():
            if n in k:
                return v
    return ""


# ── Unified row helper ───────────────────────────────────────────────────────
def blank_row(source: str) -> Dict[str, Any]:
    row = {c: "" for c in UNIFIED_COLUMNS}
    row["source"] = source
    return row


def matches_drug(row: Dict[str, Any], drug: str) -> bool:
    """Strict filter: drug must appear in an intervention/drug/title field."""
    needle = drug.lower().strip()
    if not needle:
        return True
    for f in ("drug_names", "interventions", "title", "public_title"):
        if needle in str(row.get(f, "")).lower():
            return True
    return False


# ── Excel output ─────────────────────────────────────────────────────────────
def _autosize_and_style(ws, n_cols: int):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2F5496")
    body_font = Font(name="Arial", size=10)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for i in range(1, n_cols + 1):
        letter = get_column_letter(i)
        longest = 0
        for cell in ws[letter]:
            v = cell.value
            if v is not None:
                longest = max(longest, min(len(str(v)), 60))
        ws.column_dimensions[letter].width = max(12, min(longest + 2, 60))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_excel(rows: List[Dict[str, Any]], path: str,
                columns: Optional[List[str]] = None,
                sheet_name: str = "Trials") -> Optional[str]:
    """Write rows to .xlsx. Column order: `columns` first, then any extras."""
    import pandas as pd

    if not rows:
        print(f"  (no rows – {path} not written)", file=sys.stderr)
        return None

    ordered: List[str] = list(columns) if columns else []
    for row in rows:
        for k in row:
            if k not in ordered:
                ordered.append(k)

    df = pd.DataFrame(rows).reindex(columns=ordered)
    df = df.fillna("")
    # Excel cells cap at 32767 chars
    for col in df.columns:
        df[col] = df[col].astype(str).str.slice(0, 32000)

    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
        _autosize_and_style(writer.sheets[sheet_name[:31]], len(df.columns))
    print(f"  wrote {len(df)} rows -> {path}", file=sys.stderr)
    return path


def run_cli(fetch_fn, source_name: str, description: str):
    """Shared CLI wrapper so every registry script behaves identically."""
    import argparse
    ap = argparse.ArgumentParser(description=description)
    ap.add_argument("drug", help="drug / intervention name")
    ap.add_argument("--max-records", type=int, default=None)
    ap.add_argument("--strict", action="store_true",
                    help="keep only trials naming the drug in an intervention field")
    ap.add_argument("--no-details", action="store_true",
                    help="skip per-trial detail requests (faster, fewer fields)")
    ap.add_argument("--out", default=None, help="output .xlsx path")
    args = ap.parse_args()

    slug = re.sub(r"[^a-z0-9]+", "_", args.drug.lower()).strip("_")
    out = args.out or f"{slug}_{source_name.lower().replace(' ', '_').replace('.', '')}.xlsx"

    print(f"[{source_name}] searching for '{args.drug}' ...", file=sys.stderr)
    try:
        rows = fetch_fn(args.drug, max_records=args.max_records,
                        details=not args.no_details)
    except Exception as exc:
        print(f"[{source_name}] ERROR: {exc}", file=sys.stderr)
        return 1

    if args.strict:
        rows = [r for r in rows if matches_drug(r, args.drug)]
    print(f"[{source_name}] {len(rows)} trial(s)", file=sys.stderr)
    write_excel(rows, out, UNIFIED_COLUMNS, sheet_name=source_name[:31])
    return 0
